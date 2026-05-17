"""Tests for dump_graph (FusekiStore) and GET /dump endpoint."""

from __future__ import annotations

import io
import json
from unittest.mock import AsyncMock

import pytest
from fastapi.testclient import TestClient
from rdflib import Graph, Literal, URIRef
from rdflib.namespace import RDF, RDFS


# ── helpers ───────────────────────────────────────────────────────────────────


def _make_tiny_graph() -> Graph:
    g = Graph()
    pk = URIRef("http://example.org/Pokemon")
    pikachu = URIRef("http://example.org/Pikachu")
    g.add((pikachu, RDF.type, pk))
    g.add((pikachu, RDFS.label, Literal("Pikachu")))
    return g


def _make_store(schema_graph: Graph | None = None, data_graph: Graph | None = None):
    """Return a FusekiStore with mocked _gsp_get and _ensure_dataset."""
    from ontorag.stores.fuseki import FusekiStore

    store = FusekiStore("http://localhost:3030", "test", "admin", "admin")

    schema = schema_graph if schema_graph is not None else _make_tiny_graph()
    data = data_graph if data_graph is not None else Graph()

    async def fake_gsp_get(named_graph: str) -> Graph:
        if "schema" in named_graph:
            return schema
        return data

    store._gsp_get = fake_gsp_get  # type: ignore[method-assign]
    store._ensure_dataset = AsyncMock()  # type: ignore[method-assign]
    return store


# ── FusekiStore.dump_graph ────────────────────────────────────────────────────


class TestDumpGraphTTL:
    async def test_schema_returns_turtle(self):
        store = _make_store()
        result = await store.dump_graph("schema", "ttl")
        text = result.decode()
        assert "@prefix" in text or "Pikachu" in text

    async def test_data_returns_empty_turtle_for_empty_graph(self):
        store = _make_store(data_graph=Graph())
        result = await store.dump_graph("data", "ttl")
        assert isinstance(result, bytes)

    async def test_all_merges_both_graphs(self):
        schema_g = _make_tiny_graph()
        data_g = Graph()
        extra = URIRef("http://example.org/extra")
        data_g.add((extra, RDF.type, URIRef("http://example.org/Thing")))
        store = _make_store(schema_graph=schema_g, data_graph=data_g)
        result = await store.dump_graph("all", "ttl")
        merged = Graph()
        merged.parse(data=result.decode(), format="turtle")
        assert len(merged) == len(schema_g) + len(data_g)


class TestDumpGraphJSON:
    async def test_schema_returns_json_array(self):
        store = _make_store()
        result = await store.dump_graph("schema", "json")
        rows = json.loads(result.decode())
        assert isinstance(rows, list)
        assert all("s" in r and "p" in r and "o" in r for r in rows)

    async def test_json_triple_count_matches_graph(self):
        g = _make_tiny_graph()
        store = _make_store(schema_graph=g)
        result = await store.dump_graph("schema", "json")
        rows = json.loads(result.decode())
        assert len(rows) == len(g)


class TestDumpGraphJSONL:
    async def test_schema_returns_ndjson(self):
        store = _make_store()
        result = await store.dump_graph("schema", "jsonl")
        lines = [ln for ln in result.decode().splitlines() if ln.strip()]
        for line in lines:
            obj = json.loads(line)
            assert "s" in obj and "p" in obj and "o" in obj

    async def test_jsonl_ends_with_newline(self):
        store = _make_store()
        result = await store.dump_graph("schema", "jsonl")
        assert result.endswith(b"\n")

    async def test_jsonl_empty_graph_returns_empty_bytes(self):
        store = _make_store(schema_graph=Graph())
        result = await store.dump_graph("schema", "jsonl")
        assert result == b""

    async def test_jsonl_line_count_matches_graph(self):
        g = _make_tiny_graph()
        store = _make_store(schema_graph=g)
        result = await store.dump_graph("schema", "jsonl")
        lines = [ln for ln in result.decode().splitlines() if ln.strip()]
        assert len(lines) == len(g)


class TestDumpGraphXLSX:
    async def test_schema_returns_xlsx_bytes(self):
        store = _make_store()
        result = await store.dump_graph("schema", "xlsx")
        # XLSX magic bytes: PK\x03\x04
        assert result[:4] == b"PK\x03\x04"

    async def test_xlsx_has_correct_sheet_name_for_schema(self):
        import openpyxl

        store = _make_store()
        result = await store.dump_graph("schema", "xlsx")
        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert "TBox" in wb.sheetnames

    async def test_xlsx_all_has_two_sheets(self):
        import openpyxl

        store = _make_store(
            schema_graph=_make_tiny_graph(),
            data_graph=_make_tiny_graph(),
        )
        result = await store.dump_graph("all", "xlsx")
        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert set(wb.sheetnames) == {"TBox", "ABox"}

    async def test_xlsx_row_count_matches_triples(self):
        import openpyxl

        g = _make_tiny_graph()
        store = _make_store(schema_graph=g)
        result = await store.dump_graph("schema", "xlsx")
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["TBox"]
        # first row is header
        assert ws.max_row - 1 == len(g)

    async def test_xlsx_header_row(self):
        import openpyxl

        store = _make_store()
        result = await store.dump_graph("schema", "xlsx")
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["TBox"]
        headers = [ws.cell(1, c).value for c in range(1, 4)]
        assert headers == ["Subject", "Predicate", "Object"]


# ── GET /dump endpoint ────────────────────────────────────────────────────────


@pytest.fixture()
def client():
    from ontorag.api.main import app
    from ontorag.api.deps import get_store

    mock_store = _make_store()
    app.dependency_overrides[get_store] = lambda: mock_store
    with TestClient(app) as c:
        yield c
    app.dependency_overrides.clear()


class TestDumpEndpoint:
    def test_schema_ttl_download(self, client):
        resp = client.get("/dump?target=schema&format=ttl")
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith("text/turtle")
        assert "attachment" in resp.headers["content-disposition"]
        assert resp.content

    def test_schema_json_download(self, client):
        resp = client.get("/dump?target=schema&format=json")
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith("application/json")
        rows = resp.json()
        assert isinstance(rows, list)

    def test_schema_jsonl_download(self, client):
        resp = client.get("/dump?target=schema&format=jsonl")
        assert resp.status_code == 200
        assert "ndjson" in resp.headers["content-type"]
        lines = [ln for ln in resp.text.splitlines() if ln.strip()]
        assert len(lines) > 0

    def test_schema_xlsx_download(self, client):
        resp = client.get("/dump?target=schema&format=xlsx")
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]
        assert resp.content[:4] == b"PK\x03\x04"

    def test_data_target(self, client):
        resp = client.get("/dump?target=data&format=ttl")
        assert resp.status_code == 200

    def test_all_target_xlsx_two_sheets(self, client):
        import openpyxl

        resp = client.get("/dump?target=all&format=xlsx")
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        assert "TBox" in wb.sheetnames
        assert "ABox" in wb.sheetnames

    def test_filename_contains_target_and_format(self, client):
        resp = client.get("/dump?target=schema&format=json")
        cd = resp.headers["content-disposition"]
        assert "ontorag_schema_" in cd
        assert ".json" in cd

    def test_invalid_format_returns_422(self, client):
        resp = client.get("/dump?target=schema&format=csv")
        assert resp.status_code == 422

    def test_invalid_target_returns_422(self, client):
        resp = client.get("/dump?target=tbox&format=ttl")
        assert resp.status_code == 422
